"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import subprocess
import sys
import tempfile
from pathlib import Path

from office.soffice import get_soffice_env

MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
 script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def setup_libreoffice_macro(profile_dir: Path) -> bool:
    macro_dir = profile_dir / "user/basic/Standard"
    macro_file = macro_dir / MACRO_FILENAME

    if macro_file.exists() and "RecalculateAndSave" in macro_file.read_text():
        return True

    if not macro_dir.exists():
        try:
            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                    "--terminate_after_init",
                ],
                capture_output=True,
                timeout=10,
                check=False,
                env=get_soffice_env(),
            )
        except FileNotFoundError, subprocess.TimeoutExpired:
            return False
        macro_dir.mkdir(parents=True, exist_ok=True)

    try:
        macro_file.write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    with tempfile.TemporaryDirectory(prefix="shanforge-xlsx-") as directory:
        profile_dir = Path(directory) / "libreoffice"
        if not setup_libreoffice_macro(profile_dir):
            return {"error": "Failed to setup isolated LibreOffice macro"}

        cmd = [
            "soffice",
            "--headless",
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--norestore",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False,
                env=get_soffice_env(),
            )
        except subprocess.TimeoutExpired:
            return {"error": f"LibreOffice recalculation timed out after {timeout}s"}

    if result.returncode != 0:
        error_msg = result.stderr or "Unknown error during recalculation"
        return {"error": error_msg}

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))
    if "error" in result:
        sys.exit(1)


if __name__ == "__main__":
    main()
